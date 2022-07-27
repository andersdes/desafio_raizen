import os
import urllib.request
import pandas as pd
import numpy as np

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime 

def _download_datasets():
    """
    The federal government makes available a set of public datasets, including those that will be part of this analysis.
    Link: https://dados.gov.br/dataset

    
    The datasets that are part of this analysis can be found in the links below::   
        -> https://dados.gov.br/dataset/vendas-de-derivados-de-petroleo-e-biocombustiveis/resource/c3d6e0b4-f86e-48f8-9325-6cc0d434b33f
        -> https://dados.gov.br/dataset/vendas-de-derivados-de-petroleo-e-biocombustiveis/resource/2429fdeb-df86-4e63-b248-2038f6c3e3cc
    """ 
        
    # Download Files
    try:
        # Path the storage 
        path = os.path.dirname(os.path.abspath(__file__)) + '/dados/'
        
        # Create paht case not exist
        if not os.path.exists(path):
            os.makedirs(path)

        print('********Start - Download Datasets********')
        # Link reference            
        link_ref = 'https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-abertos/arquivos/vdpb'
        
        # Download - Sales of oil derivative fuels by UF and product        
        url = f'{link_ref}/vendas-derivados-petroleo-e-etanol/vendas-derivados-petroleo-etanol-m3-1990-2022.csv'
        urllib.request.urlretrieve(url, path + "dataset_derivative.csv")
        print("Download successful - Derivative Fuels")  

        # Download - Sales of oil derivative fuels by UF and product        
        url = f'{link_ref}/vct/vendas-oleo-diesel-tipo-m3-2013-2022.csv'
        urllib.request.urlretrieve(url, path + "dataset_diesel.csv")
        print("Download successful - Diesel")          
        print('********End - Download Datasets********', end='\n\n')
    except:
        print("Download Dataset failed")
        print('********End - Download Datasets********', end='\n\n')

def _download_data_pivot():
    """
    Analysis file (pivoted data)
        File provided by Raizen
        -> https://github.com/raizen-analytics/data-engineering-test/raw/master/assets/vendas-combustiveis-m3.xls
        File provided by Federal government
        XLX
            -> https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-estatisticos/de/vdpb/vendas-combustiveis-m3.xls
        XLSX
        -> https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-estatisticos/de/vdpb/vendas-combustiveis-m3.xls/view
            -> permite baixar o xlsx através do link :
            https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-estatisticos/de/vdpb/vendas-combustiveis-m3.xls/@@download/file/vendas-combustiveis-m3.xlsx
    """
    try:
        # Path the storage 
        path = os.path.dirname(os.path.abspath(__file__)) + '/dados/'
        print('********Start - Download Pivot********')
        # Download - Sales of oil derivative fuels by UF and product        
        url = 'https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-estatisticos/de/vdpb/vendas-combustiveis-m3.xls/@@download/file/vendas-combustiveis-m3.xlsx'
        urllib.request.urlretrieve(url, path + "vendas-combustiveis-m3.xlsx")
    
        print("Download successful - Data Pivot") 
        print('********End - Download Pivot********', end='\n\n')
    except:
        print("Download Dataset failed")
        print('********End - Download Pivot********', end='\n\n')


def formated_year_month(year, month):
    """
    Format field year_month	date

    Parameters
    ----------
    year : String
        Year extraction from pivot table.
    month : String
        Month extraction from pivot table.
        
    Returns
    -------
    date : String
        Returns returns the date according to the past period.
    """  
    month_name = {
        'JAN':1,
        'FEV':2,
        'MAR':3,
        'ABR':4,
        'MAI':5,
        'JUN':6,
        'JUL':7,
        'AGO':8,
        'SET':9,
        'OUT':10,
        'NOV':11,
        'DEZ':12        
    }
    date = datetime(int(year), month_name[month], 1)
    #date = f'{str(year)}_{str(month_name[month])}'
    return date

def trim_all_columns(dataframe):
    """
    Trim whitespace from ends of each value across all series in dataframe
    
    Parameters
    ----------
    dataframe : array
        Data extracted from excel file.            
        
    Returns:
    --------
        Dataframe Cleaning
    """

    trim_strings = lambda x: x.strip() if isinstance(x, str) else x
    return dataframe.applymap(trim_strings)

def get_total_pivot(df):
    """
    Calculates the pivot total.

    Parameters
    ----------
    df : ndarray
        Data Frame.

    Returns
    -------
    df : ndarray
        Returns a dataframe with the consolidated total by year.
    """         
    # Filter row with values total
    df = df.query("Mês == 'Total do Ano'")   
    # It's used to create a specific format of the DataFrame object where one or more columns work as identifiers.
    df = df.melt(id_vars=["Mês"], var_name="year_total", value_name="volume_total")
    df['year_total'] = df['year_total'].astype(int)
    # Value format
    df['volume_total'] = df['volume_total'].apply(lambda x:  "{:.2f}".format(x))     
    # sorting dataframe
    df = df.sort_values(by=['year_total'])
    # Drop columns    
    df = df.drop(columns=['Mês']);
    
    return df

def get_total_dataframe(df):
    """
    Calculates the dataframe total.

    Parameters
    ----------
    df : ndarray
        Data Frame.

    Returns
    -------
    df : ndarray
        Returns a dataframe with the consolidated total by year.
    """      
    # Drop columns
    df = df.drop(columns=['uf','product','unit','created_at']);
    # It's used to create a specific format of the DataFrame object where one or more columns work as identifiers.
    df = df.melt(id_vars=["year_month"], var_name="vol", value_name="volume_df")
    # Get year
    df['year_df'] = df['year_month'].apply(lambda x: x[:4]).astype(int)
    # Drop columns
    df = df.drop(columns=['vol', 'year_month']);
    # Group by year and sum column volume
    df = df.groupby(['year_df'])['volume_df'].sum().reset_index()
    # Value format
    df['volume_df'] = df['volume_df'].apply(lambda x:  "{:.2f}".format(x))  
    return df

def clean_dataframe(file_name, start_period):
    """
    Clean up the dataframe and name the columns.

    Parameters
    ----------
    df : ndarray
        Dataframe.
    start_period: int
        Period you want to return from the dataset.   

    Returns
    -------
    df : ndarray
        Dataframe.
    """  
    # Path dataset
    path = os.path.dirname(os.path.abspath(__file__)) + '/dados/' + file_name + '.csv'
    
    print('********Start - Data Clean********')
    # Mount dataframe
    df = pd.read_csv(
        path,
        index_col=None,
        delimiter=';'
    )
    # Set titles columns
    header_df = ['year', 'month', 'region', 'uf', 'product', 'volume']
    df.columns = header_df
    
    # sorting dataframe
    df = df.sort_values(by=['year', 'month', 'uf', 'product'])

    # Filter Period
    df = df.where(df['year']>=int(start_period))
        
    # Data cleaning        
    print(f'Data Clean - {file_name.upper()}')
    df = df.dropna(how='all')
    df = df.fillna(np.nan).replace([np.nan], [None])
    df = trim_all_columns(df)
    
    # Data transformation
    df['year_month'] = df.apply(lambda x: formated_year_month(int(x['year']), x['month']), axis=1)
    df['unit'] = 'm3'
    df['volume'] = df['volume'].str.replace(',', '.').apply(lambda x: float(x))
    df['created_at'] = pd.Timestamp.now().strftime('%Y-%m-%d %X')
    
    # Drop columns    
    df = df.drop(columns=['year','month', 'region']);

    # Reordering columns
    df = df[['year_month', 'uf', 'product', 'unit', 'volume', 'created_at']]
    # Order by columns
    df = df.sort_values(by=['year_month', 'uf', 'product'])
    
    # File csv
    df.to_csv(path, sep = ';', index=False)
    print(f'Generated dataset - {file_name.upper()}')
    
    print('********End - Data Clean********', end='\n\n')
    return df

def _clean_file():
    """
    This function is intended to carry out the cleaning process of downloaded datasets.
    """  
    clean_dataframe('dataset_derivative', 2000)
    clean_dataframe('dataset_diesel', 2013)

def _generation_file():
    """
    This function generates the final file with the result of the extracted datasets in addition to consolidating these datasets.
    """
    
    # Paths files
    path = os.path.dirname(os.path.abspath(__file__)) + '/dados/'
    file_derivative = path + 'dataset_derivative.csv'
    file_diesel = path + 'dataset_diesel.csv'
    file_final = path + 'data_extracted.xlsx'
    
    df_deravative = pd.read_csv(file_derivative,delimiter=';')
    df_diesel = pd.read_csv(file_diesel,delimiter=';')
    df_final = pd.concat([df_deravative, df_diesel], ignore_index=True, sort=False)
    wb = Workbook()
    
    print('********Start - Create File Final********')
    
    # Insert row derivative
    ws1 = wb.create_sheet('DERIVATIVES',0)
    for row in dataframe_to_rows(df_deravative, index = False):
        ws1.append(row)
    print('Create Sheet - DERIVATIVES')
           
    # Insert row diesel
    ws2 = wb.create_sheet('DIESEL',1)
    for row in dataframe_to_rows(df_diesel, index = False):
        ws2.append(row)
    print('Create Sheet - DIESEL')        
        
    # Insert row final
    ws3 = wb.create_sheet('DERIVATIVES_DISEL_FINAL',2)
    for row in dataframe_to_rows(df_final, index = False):
        ws3.append(row)
    print('Create Sheet - DERIVATIVES_DISEL_FINAL')           
    
    # Remove existing sheet template
    wb.remove(wb['Sheet'])
    
    # Save file final
    wb.save(file_final)      
    print('********End - Create File Final********', end='\n\n')


def _check_results():
    """
    This function checks if the result extracted from the datasets match the data in the pivoted file.
    """
    print('********Start - Check Result********')
    # Paths files
    path = os.path.dirname(os.path.abspath(__file__)) + '/dados/'
    file_derivative = path + 'dataset_derivative.csv'
    file_diesel = path + 'dataset_diesel.csv'
    file_pivot = path + 'vendas-combustiveis-m3.xlsx'
    file_result = path + 'data_extracted.xlsx'
    
    # Load file final
    wb = load_workbook(filename = file_result)
    
    
    # Compare total Derivative x Pivot
    # Extract cvs Derivative
    df_deravative = pd.read_csv(file_derivative,delimiter=';')
    
    # Extract pivot derivative
    df_pivot_derivative = pd.read_excel(
                    io=file_pivot,
                    sheet_name="Plan1",
                    index_col=None,
                    header=52,
                    nrows = 13,
                    na_values=["NA"],
                    usecols="B:Y",
                    engine="openpyxl",
                )
    
    df_result_derivative = pd.concat([get_total_pivot(df_pivot_derivative), get_total_dataframe(df_deravative)], axis=1, join="inner")
    df_result_derivative['value_equal'] = df_result_derivative['volume_total'].equals(df_result_derivative['volume_df'])
    
    ws1 = wb.create_sheet('RESULT_DERIVATIVESxPIVOT',3)
     # Insert row final
    for row in dataframe_to_rows(df_result_derivative, index = False):
        ws1.append(row)
    
    print('Check Result - RESULT_DERIVATIVESxPIVOT')
    
    # Compare total Diesel x Pivot
    # Extract cvs Diesel
    df_diesel = pd.read_csv(file_diesel,delimiter=';')
    
    # Extract pivot derivative
    df_pivot_diesel = pd.read_excel(
                    io=file_pivot,
                    sheet_name="Plan1",
                    index_col=None,
                    header=188,
                    nrows = 13,
                    na_values=["NA"],
                    usecols="B:L",
                    engine="openpyxl",
                )
    
    df_result_diesel = pd.concat([get_total_pivot(df_pivot_diesel), get_total_dataframe(df_diesel)], axis=1, join="inner")
    df_result_diesel['value_equal'] = df_result_diesel['volume_total'].equals(df_result_diesel['volume_df'])
    
    ws2 = wb.create_sheet('RESULT_DIESELxPIVOT',4)
     # Insert row final
    for row in dataframe_to_rows(df_result_diesel, index = False):
        ws2.append(row)
    
    print('Check Result - RESULT_DIESELxPIVOT')
           
    # Save file final
    wb.save(file_result)
    print('********End - Check Result********')

'''
Execution Sequence
_download_datasets()
_download_data_pivot()
_clean_file()
_generation_file()
_check_results()
'''
